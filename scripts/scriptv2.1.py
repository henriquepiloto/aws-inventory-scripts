import boto3
import json
import datetime
import pandas as pd
from botocore.exceptions import ClientError
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

BUCKET_NAME = "script-piloto"
ROLES_KEY = "tsj.json"
output_prefix = "relatorios/"
output_local_prefix = "/tmp/"

# -------------------------------
# Utilitários
# -------------------------------
def get_roles_from_s3():
    s3 = boto3.client("s3")
    obj = s3.get_object(Bucket=BUCKET_NAME, Key=ROLES_KEY)
    return json.loads(obj['Body'].read())

def assume_role(account_id, role_name):
    sts = boto3.client("sts")
    response = sts.assume_role(
        RoleArn=f"arn:aws:iam::{account_id}:role/{role_name}",
        RoleSessionName="AuditSession"
    )
    return response['Credentials']

def make_creds_dict(creds):
    return {
        "aws_access_key_id": creds["AccessKeyId"],
        "aws_secret_access_key": creds["SecretAccessKey"],
        "aws_session_token": creds["SessionToken"]
    }

# -------------------------------
# IAM Checks
# -------------------------------
def check_iam(creds, cliente, account_id):
    iam = boto3.client("iam", **creds)
    findings = []
    try:
        users = iam.list_users()["Users"]
    except Exception:
        users = []

    for user in users:
        uname = user["UserName"]
        pwd_used = user.get("PasswordLastUsed")
        has_console_access = pwd_used is not None

        mfa_devices = iam.list_mfa_devices(UserName=uname)["MFADevices"]
        if has_console_access and not mfa_devices:
            findings.append({
                "Conta": cliente, "ID da Conta": account_id,
                "Serviço": "IAM", "Descrição": f"Usuário IAM sem MFA: {uname}",
                "Motivo": "Usuários IAM com acesso à console devem ter MFA habilitado.",
                "Severidade": "Alta", "Recomendacao de solucao": "Ative o MFA para o usuário no IAM."
            })

        keys = iam.list_access_keys(UserName=uname)["AccessKeyMetadata"]
        for key in keys:
            age = (datetime.datetime.now(datetime.timezone.utc) - key["CreateDate"]).days
            if age > 90:
                findings.append({
                    "Conta": cliente, "ID da Conta": account_id,
                    "Serviço": "IAM", "Descrição": f"Chave de acesso antiga (>90 dias): {uname}",
                    "Motivo": "Chaves de acesso antigas aumentam o risco de comprometimento.",
                    "Severidade": "Média", "Recomendacao de solucao": "Rotacione as chaves e implemente política de expiração."
                })

    roles = iam.list_roles()["Roles"]
    for role in roles:
        name = role["RoleName"]
        try:
            policies = iam.list_attached_role_policies(RoleName=name)["AttachedPolicies"]
            for p in policies:
                policy = iam.get_policy(PolicyArn=p["PolicyArn"])["Policy"]
                version = iam.get_policy_version(
                    PolicyArn=p["PolicyArn"], VersionId=policy["DefaultVersionId"]
                )
                stmts = version["PolicyVersion"]["Document"].get("Statement", [])
                if not isinstance(stmts, list):
                    stmts = [stmts]
                for stmt in stmts:
                    if "*" in str(stmt.get("Action")) or "*" in str(stmt.get("Resource")):
                        findings.append({
                            "Conta": cliente, "ID da Conta": account_id,
                            "Serviço": "IAM", "Descrição": f"Role com permissões amplas (*): {name}",
                            "Motivo": "Permissões com curingas (*) permitem acesso irrestrito.",
                            "Severidade": "Alta", "Recomendacao de solucao": "Restringir permissões da role."
                        })
        except Exception:
            continue
    return findings

# -------------------------------
# S3 Checks
# -------------------------------
def check_s3(creds, cliente, account_id):
    s3 = boto3.client("s3", **creds)
    findings = []
    try:
        buckets = s3.list_buckets()["Buckets"]
    except Exception:
        return findings

    for b in buckets:
        bucket_name = b["Name"]
        try:
            acl = s3.get_bucket_acl(Bucket=bucket_name)
            for grant in acl['Grants']:
                if grant.get('Grantee', {}).get('URI') == 'http://acs.amazonaws.com/groups/global/AllUsers':
                    findings.append({
                        "Conta": cliente, "ID da Conta": account_id,
                        "Serviço": "S3", "Descrição": f"Bucket público: {bucket_name}",
                        "Motivo": "Acesso público a dados pode gerar vazamentos.",
                        "Severidade": "Alta", "Recomendacao de solucao": "Ativar Block Public Access e revisar ACLs."
                    })
        except Exception:
            continue
        try:
            s3.get_bucket_encryption(Bucket=bucket_name)
        except ClientError:
            findings.append({
                "Conta": cliente, "ID da Conta": account_id,
                "Serviço": "S3", "Descrição": f"Bucket sem criptografia: {bucket_name}",
                "Motivo": "Dados não criptografados em repouso podem ser acessados indevidamente.",
                "Severidade": "Média", "Recomendacao de solucao": "Ativar criptografia SSE-KMS no bucket."
            })
    return findings

# -------------------------------
# Security Groups Checks
# -------------------------------
def check_security_groups(creds, cliente, account_id):
    ec2 = boto3.client("ec2", **creds)
    findings = []
    try:
        sgs = ec2.describe_security_groups()["SecurityGroups"]
    except Exception:
        return findings

    for sg in sgs:
        for rule in sg.get("IpPermissions", []):
            from_port = rule.get("FromPort")
            for ip_range in rule.get("IpRanges", []):
                if ip_range.get("CidrIp") == "0.0.0.0/0":
                    findings.append({
                        "Conta": cliente, "ID da Conta": account_id,
                        "Serviço": "EC2", "Descrição": f"SG {sg['GroupName']} aberto na porta {from_port}",
                        "Motivo": "Exposição pública a todas as IPs.",
                        "Severidade": "Alta", "Recomendacao de solucao": "Restringir IPs de acesso."
                    })
    return findings

# -------------------------------
# RDS Checks
# -------------------------------
def check_rds(creds, cliente, account_id):
    rds = boto3.client("rds", **creds)
    findings = []
    try:
        instances = rds.describe_db_instances()["DBInstances"]
    except Exception:
        return findings

    for db in instances:
        db_id = db["DBInstanceIdentifier"]

        if db["PubliclyAccessible"]:
            findings.append({
                "Conta": cliente, "ID da Conta": account_id,
                "Serviço": "RDS", "Descrição": f"RDS público: {db_id}",
                "Motivo": "Instância exposta publicamente.",
                "Severidade": "Alta", "Recomendacao de solucao": "Tornar o RDS privado."
            })

        if not db.get("StorageEncrypted", False):
            findings.append({
                "Conta": cliente, "ID da Conta": account_id,
                "Serviço": "RDS", "Descrição": f"RDS sem criptografia: {db_id}",
                "Motivo": "Dados não criptografados.",
                "Severidade": "Média", "Recomendacao de solucao": "Ativar criptografia KMS."
            })
    return findings

# -------------------------------
# EBS Checks
# -------------------------------
def check_ebs(creds, cliente, account_id):
    ec2 = boto3.client("ec2", **creds)
    findings = []
    try:
        volumes = ec2.describe_volumes()["Volumes"]
    except Exception:
        return findings

    for vol in volumes:
        vol_id = vol["VolumeId"]
        if not vol.get("Encrypted", False):
            findings.append({
                "Conta": cliente, "ID da Conta": account_id,
                "Serviço": "EBS", "Descrição": f"Volume EBS sem criptografia: {vol_id}",
                "Motivo": "Volume não criptografado pode expor dados.",
                "Severidade": "Média", "Recomendacao de solucao": "Criar snapshots criptografados."
            })
    return findings

# -------------------------------
# KMS Checks
# -------------------------------
def check_kms(creds, cliente, account_id):
    kms = boto3.client("kms", **creds)
    findings = []
    try:
        keys = kms.list_keys()["Keys"]
    except Exception:
        return findings

    for key_info in keys:
        key_id = key_info["KeyId"]
        meta = kms.describe_key(KeyId=key_id)["KeyMetadata"]
        if not meta.get("KeyRotationEnabled", False):
            findings.append({
                "Conta": cliente, "ID da Conta": account_id,
                "Serviço": "KMS", "Descrição": f"Chave KMS sem rotação: {meta['KeyId']}",
                "Motivo": "Chaves sem rotação são mais vulneráveis.",
                "Severidade": "Baixa", "Recomendacao de solucao": "Ativar KeyRotation."
            })
    return findings

# -------------------------------
# WAF Checks
# -------------------------------
def check_waf(creds, cliente, account_id):
    waf = boto3.client("wafv2", **creds, region_name="us-east-1")
    findings = []
    try:
        web_acls = waf.list_web_acls(Scope='REGIONAL')['WebACLs']
    except Exception:
        return findings

    for acl in web_acls:
        if acl.get("Description", "") == "":
            findings.append({
                "Conta": cliente, "ID da Conta": account_id,
                "Serviço": "WAF", "Descrição": f"WebACL sem descrição: {acl['Name']}",
                "Motivo": "Pode estar sem documentação.",
                "Severidade": "Baixa", "Recomendacao de solucao": "Adicione descrição às ACLs."
            })
    return findings

# -------------------------------
# Pontuação e Excel
# -------------------------------
def calcular_pontuacao(relatorio):
    score = 100
    pesos = {"Alta": 5, "Média": 3, "Baixa": 1}
    for item in relatorio:
        score -= pesos.get(item.get("Severidade"), 0)
    return max(score, 0)

def format_excel_with_resumo(file_path, relatorio, score):
    df = pd.DataFrame(relatorio)

    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="Achados", index=False)
        resumo_df = pd.DataFrame({
            "Severidade": ["Alta", "Média", "Baixa"],
            "Quantidade": [
                df[df["Severidade"] == "Alta"].shape[0],
                df[df["Severidade"] == "Média"].shape[0],
                df[df["Severidade"] == "Baixa"].shape[0],
            ]
        })
        resumo_df.loc[len(resumo_df)] = ["Pontuação Final", score]
        resumo_df.to_excel(writer, sheet_name="Resumo", index=False)

    wb = load_workbook(file_path)
    ws = wb["Achados"]
    severity_fill = {
        "Alta": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "Média": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "Baixa": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    }
    for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
        severity = row[5].value
        if severity in severity_fill:
            for cell in row:
                cell.fill = severity_fill[severity]
    wb.save(file_path)

# -------------------------------
# MAIN
# -------------------------------
def main():
    relatorio = []
    roles = get_roles_from_s3()

    for entry in roles:
        account_id = entry["account_id"]
        role_name = entry["role_name"]
        cliente = entry.get("cliente", "Desconhecido")

        creds_raw = assume_role(account_id, role_name)
        creds = make_creds_dict(creds_raw)

        print(f"[Verificando conta {cliente} ({account_id})]")
        relatorio.extend(check_iam(creds, cliente, account_id))
        relatorio.extend(check_s3(creds, cliente, account_id))
        relatorio.extend(check_security_groups(creds, cliente, account_id))
        relatorio.extend(check_rds(creds, cliente, account_id))
        relatorio.extend(check_ebs(creds, cliente, account_id))
        relatorio.extend(check_kms(creds, cliente, account_id))
        relatorio.extend(check_waf(creds, cliente, account_id))

    if relatorio:
        score = calcular_pontuacao(relatorio)
        print(f"[✓] Pontuação de segurança: {score}/100")

        timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M")
        filename = f"relatorio_seg_{timestamp}.xlsx"
        format_excel_with_resumo(filename, relatorio, score)
        print(f"[✓] Relatório salvo: {filename}")

        s3_client = boto3.client("s3")
        try:
            s3_key = output_prefix + filename
            s3_client.upload_file(filename, BUCKET_NAME, s3_key)
            print(f"[✓] Upload para o S3 realizado: s3://{BUCKET_NAME}/{s3_key}")
        except Exception as e:
            print(f"[x] Erro ao enviar para o S3: {e}")
    else:
        print("[INFO] Nenhum item de segurança encontrado.")

if __name__ == "__main__":
    main()
